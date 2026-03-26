"""Parse a Bank Discount transaction history export (xlsx) into a list of dicts.

The export is obtained via: פעולות בחשבון → Excel export from the Bank Discount
brokerage portal. Layout assumptions:
  - Header row: contains the text "שם נייר" in one of its cells
  - Transaction rows: immediately follow the header row, one transaction per row
  - Columns per transaction row:
      A → security_name
      B → security_id
      C → transaction_type  (e.g. "קניה", "מכירה")
      D → quantity          (stored as positive; absolute value taken)
      E → execution_price
      F → value_ils
      G → execution_date    (datetime cell)
      H → security_type
      I → net_value_ils
      J → tax_amount
      K → commission
      N → value_date        (datetime cell)
"""

import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

logger = logging.getLogger(__name__)

# openpyxl column indices (1-based)
_COL_SEC_NAME = 1       # A — security name
_COL_SEC_ID = 2         # B — security ID (TASE numeric)
_COL_TX_TYPE = 3        # C — transaction type
_COL_QUANTITY = 4       # D — quantity
_COL_EXEC_PRICE = 5     # E — execution price
_COL_VALUE = 6          # F — gross value ILS
_COL_EXEC_DATE = 7      # G — execution date
_COL_SEC_TYPE = 8       # H — security type
_COL_NET_VALUE = 9      # I — net value ILS
_COL_TAX = 10           # J — tax amount
_COL_COMMISSION = 11    # K — commission
_COL_VALUE_DATE = 14    # N — value date


def _to_decimal(value: Any, label: str) -> Decimal | None:
    """Convert a cell value to Decimal, returning None on failure.

    Args:
        value: Raw openpyxl cell value (int, float, str, or None).
        label: Human-readable label used in log messages.

    Returns:
        Decimal representation of the value, or None if conversion fails.
    """
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except InvalidOperation:
        logger.warning("transaction_parser: cannot convert %r to Decimal for '%s'", value, label)
        return None


def _to_datetime(value: Any, label: str) -> datetime | None:
    """Extract a datetime from an openpyxl cell value.

    openpyxl returns datetime cells as Python datetime objects directly.
    String fallback handles unexpected text-formatted date cells.

    Args:
        value: Raw openpyxl cell value.
        label: Human-readable label used in log messages.

    Returns:
        datetime object, or None if conversion fails.
    """
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    # Some exports store dates as strings
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(str(value).strip(), fmt)
        except ValueError:
            continue
    logger.warning("transaction_parser: cannot parse datetime %r for '%s'", value, label)
    return None


def parse_transaction_history(filepath: str) -> list[dict]:
    """Parse a Bank Discount transaction history xlsx export.

    Opens the workbook, locates the header row by searching for a cell
    containing "שם נייר", then reads every subsequent row as a transaction
    until the end of data. Rows missing both security_id and security_name
    are skipped.

    Args:
        filepath: Absolute or relative path to the xlsx file.

    Returns:
        List of transaction dicts sorted by execution_date ascending.
        Each dict has keys: security_name, security_id, transaction_type,
        quantity, execution_price, value_ils, execution_date, security_type,
        net_value_ils, tax_amount, commission, value_date.

    Raises:
        FileNotFoundError: If the xlsx file does not exist.
        ValueError: If the header row containing "שם נייר" cannot be found.
        ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to parse Bank Discount exports. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ── Locate header row ─────────────────────────────────────────────────────
    header_row: int | None = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "שם נייר" in cell.value:
                header_row = cell.row
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(
            "transaction_parser: could not find header row — "
            "no cell containing 'שם נייר' found in the workbook"
        )

    # ── Read transaction rows ─────────────────────────────────────────────────
    transactions: list[dict] = []

    for row_idx in range(header_row + 1, ws.max_row + 1):
        sec_name = ws.cell(row_idx, _COL_SEC_NAME).value
        sec_id = ws.cell(row_idx, _COL_SEC_ID).value

        # Skip completely empty rows
        if sec_name is None and sec_id is None:
            continue

        tx_type_raw = ws.cell(row_idx, _COL_TX_TYPE).value
        tx_type = str(tx_type_raw).strip() if tx_type_raw is not None else ""

        qty_raw = _to_decimal(ws.cell(row_idx, _COL_QUANTITY).value, f"row{row_idx}.quantity")
        quantity = abs(qty_raw) if qty_raw is not None else None

        transactions.append({
            "security_name":  str(sec_name).strip() if sec_name is not None else "",
            "security_id":    str(sec_id).strip() if sec_id is not None else "",
            "transaction_type": tx_type,
            "quantity":       quantity,
            "execution_price": _to_decimal(
                ws.cell(row_idx, _COL_EXEC_PRICE).value, f"row{row_idx}.execution_price"
            ),
            "value_ils":      _to_decimal(
                ws.cell(row_idx, _COL_VALUE).value, f"row{row_idx}.value_ils"
            ),
            "execution_date": _to_datetime(
                ws.cell(row_idx, _COL_EXEC_DATE).value, f"row{row_idx}.execution_date"
            ),
            "security_type":  str(ws.cell(row_idx, _COL_SEC_TYPE).value or "").strip(),
            "net_value_ils":  _to_decimal(
                ws.cell(row_idx, _COL_NET_VALUE).value, f"row{row_idx}.net_value_ils"
            ),
            "tax_amount":     _to_decimal(
                ws.cell(row_idx, _COL_TAX).value, f"row{row_idx}.tax_amount"
            ),
            "commission":     _to_decimal(
                ws.cell(row_idx, _COL_COMMISSION).value, f"row{row_idx}.commission"
            ),
            "value_date":     _to_datetime(
                ws.cell(row_idx, _COL_VALUE_DATE).value, f"row{row_idx}.value_date"
            ),
        })

    # Sort chronologically; rows with no date sort to the front
    transactions.sort(key=lambda t: t["execution_date"] or datetime.min)

    logger.info("transaction_parser: parsed %d transactions from %s", len(transactions), filepath)
    return transactions


def format_transactions_for_prompt(transactions: list[dict]) -> str:
    """Format a transaction list as a concise text block for Claude's prompt.

    Produces one line per transaction in the format:
        YYYY-MM-DD | <type> | <name> (<id>) | <qty> units @ ₪<price> | Total ₪<net>

    Args:
        transactions: List of transaction dicts as returned by
            parse_transaction_history(), sorted oldest-first.

    Returns:
        Multi-line string suitable for inclusion in a Claude prompt as a
        <transaction_history> block. Returns an empty string if the list
        is empty.
    """
    if not transactions:
        return ""

    lines = ["TRANSACTION HISTORY (oldest to newest):"]
    for tx in transactions:
        date_str = tx["execution_date"].strftime("%Y-%m-%d") if tx["execution_date"] else "unknown"
        name_id = f"{tx['security_name']} ({tx['security_id']})" if tx["security_id"] else tx["security_name"]
        qty_str = f"{float(tx['quantity']):,.0f}" if tx["quantity"] is not None else "?"
        price_str = f"₪{float(tx['execution_price']):,.2f}" if tx["execution_price"] is not None else "₪?"
        net = tx["net_value_ils"] or tx["value_ils"]
        net_str = f"₪{float(net):,.0f}" if net is not None else "₪?"
        lines.append(
            f"{date_str} | {tx['transaction_type']} | {name_id} | "
            f"{qty_str} units @ {price_str} | Total {net_str}"
        )

    return "\n".join(lines)
