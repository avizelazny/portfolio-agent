"""Parse a Bank Discount portfolio export (xlsx) into a PortfolioSnapshot.

The export format is a fixed-layout Hebrew xlsx produced by the Bank Discount
brokerage portal. Key layout assumptions:
  - Row 4:        portfolio summary — total value (col A), day P&L ILS (col G),
                  day P&L % (col H, as a decimal fraction e.g. 0.012 = 1.2%)
  - Header row:   contains the text "שם נייר" (security name) in column A
  - Holding rows: immediately follow the header row, one security per row
  - Total row:    first row after holdings whose column A contains "סה\"כ"
  - Columns per holding row:
      A → company_name
      B → ticker (TASE numeric ID, stored as-is)
      C → market_value_ils
      D → weight_pct (decimal fraction, e.g. 0.15 = 15%)
      E → quantity
      G → current_price
      H → unrealized_pnl_pct (decimal fraction)
      K → avg_cost_ils
"""
import logging
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any

from src.models.market import Holding, PortfolioSnapshot

logger = logging.getLogger(__name__)

# openpyxl column indices (1-based)
_COL_NAME = 1       # A — company name
_COL_TICKER = 2     # B — TASE security ID
_COL_MV = 3         # C — market value ILS
_COL_WEIGHT = 4     # D — weight as decimal fraction
_COL_QTY = 5        # E — quantity
_COL_PRICE = 7      # G — current price (in agorot for mutual funds)
_COL_PNL_PCT = 8    # H — unrealized P&L % as decimal fraction
_COL_AVG_COST = 11  # K — average cost ILS (in agorot for mutual funds)
_COL_SEC_TYPE = 9   # I — security type (e.g. "קרן נאמנות")
_COL_EXCHANGE = 14  # N — exchange (e.g. "ת\"א")

# Security type string that indicates mutual fund (prices in agorot, not shekels)
_SEC_TYPE_FUND = "קרן נאמנות"
_AGOROT_DIVISOR = Decimal("100")

# Summary row index (1-based)
_SUMMARY_ROW = 4


def _d(value: Any, label: str) -> Decimal:
    """Convert a cell value to Decimal, raising ValueError on failure.

    Args:
        value: Raw openpyxl cell value (int, float, str, or None).
        label: Human-readable label used in the error message.

    Returns:
        Decimal representation of the value.

    Raises:
        ValueError: If the value is None or cannot be converted.
    """
    if value is None:
        raise ValueError(f"discount_parser: expected a number for '{label}', got None")
    try:
        return Decimal(str(value))
    except InvalidOperation:
        raise ValueError(
            f"discount_parser: cannot convert {value!r} to Decimal for '{label}'"
        )


def parse_discount_export(filepath: str) -> PortfolioSnapshot:
    """Parse a Bank Discount portfolio xlsx export into a PortfolioSnapshot.

    Opens the workbook, reads the portfolio summary from row 4, locates the
    holdings table by searching for the "שם נייר" header cell, and reads each
    holding row until the "סה\"כ" totals row. Cash is inferred as the
    difference between total portfolio value and the sum of all market values.

    Args:
        filepath: Absolute or relative path to the xlsx file.

    Returns:
        A fully populated PortfolioSnapshot.

    Raises:
        FileNotFoundError: If the xlsx file does not exist.
        ValueError: If required cells are missing, malformed, or the header
            row cannot be located.
        ImportError: If openpyxl is not installed.
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required to parse Bank Discount exports. "
            "Install it with: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.active

    # ── Summary row (row 4) ────────────────────────────────────────────────────
    total_value_ils = _d(ws.cell(_SUMMARY_ROW, _COL_NAME).value, "total_value_ils")

    _day_pnl_ils_raw = ws.cell(_SUMMARY_ROW, _COL_PRICE).value
    day_pnl_ils = _d(_day_pnl_ils_raw, "day_pnl_ils") if _day_pnl_ils_raw is not None else Decimal("0")

    _day_pnl_pct_raw = ws.cell(_SUMMARY_ROW, _COL_PNL_PCT).value
    day_pnl_pct = (
        _d(_day_pnl_pct_raw, "day_pnl_pct") * Decimal("100")
        if _day_pnl_pct_raw is not None
        else Decimal("0")
    )

    # ── Locate header row ─────────────────────────────────────────────────────
    header_row: int | None = None
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and "שם נייר" in cell.value:
                header_row = cell.row
                break
        if header_row is not None:
            break

    if header_row is None:
        raise ValueError(
            "discount_parser: could not find header row — "
            "no cell containing 'שם נייר' found in the workbook"
        )

    # ── Read holdings ─────────────────────────────────────────────────────────
    holdings: list[Holding] = []
    first_data_row = header_row + 1

    for row_idx in range(first_data_row, ws.max_row + 1):
        name_cell = ws.cell(row_idx, _COL_NAME).value

        # Stop at the totals row
        if isinstance(name_cell, str) and "סה" in name_cell:
            break

        # Skip completely empty rows
        if name_cell is None:
            continue

        company_name = str(name_cell).strip()
        ticker = str(ws.cell(row_idx, _COL_TICKER).value or "").strip()
        if not ticker:
            logger.warning("discount_parser: skipping row %d — empty ticker", row_idx)
            continue

        try:
            market_value_ils = _d(ws.cell(row_idx, _COL_MV).value, f"{ticker}.market_value_ils")
            weight_pct = _d(ws.cell(row_idx, _COL_WEIGHT).value, f"{ticker}.weight_pct") * Decimal("100")
            quantity = _d(ws.cell(row_idx, _COL_QTY).value, f"{ticker}.quantity")
            current_price = _d(ws.cell(row_idx, _COL_PRICE).value, f"{ticker}.current_price")
            pnl_pct = _d(ws.cell(row_idx, _COL_PNL_PCT).value, f"{ticker}.unrealized_pnl_pct") * Decimal("100")
            avg_cost_ils = _d(ws.cell(row_idx, _COL_AVG_COST).value, f"{ticker}.avg_cost_ils")
        except ValueError as exc:
            logger.warning("discount_parser: skipping %s — %s", ticker, exc)
            continue

        # Bank Discount exports mutual fund prices in agorot (1/100 shekel).
        # Detect by security type column (col I = "קרן נאמנות") and divide by 100.
        # Collapse internal whitespace and strip before matching — export may have
        # trailing/leading spaces or double-spaces within the value.
        sec_type_raw = ws.cell(row_idx, _COL_SEC_TYPE).value or ""
        sec_type = " ".join(str(sec_type_raw).split())  # normalise all whitespace
        logger.debug("discount_parser: %s — sec_type=%r (raw=%r)", ticker, sec_type, sec_type_raw)
        if _SEC_TYPE_FUND in sec_type:
            current_price = current_price / _AGOROT_DIVISOR
            avg_cost_ils = avg_cost_ils / _AGOROT_DIVISOR
            logger.debug(
                "discount_parser: %s — converted fund prices from agorot: "
                "current=₪%.4f avg_cost=₪%.4f",
                ticker, current_price, avg_cost_ils,
            )

        unrealized_pnl_ils = market_value_ils - (quantity * avg_cost_ils)

        holdings.append(Holding(
            ticker=ticker,
            company_name=company_name,
            quantity=quantity,
            avg_cost_ils=avg_cost_ils,
            current_price=current_price,
            market_value_ils=market_value_ils,
            unrealized_pnl_ils=unrealized_pnl_ils,
            unrealized_pnl_pct=pnl_pct,
            weight_pct=weight_pct,
        ))

    if not holdings:
        raise ValueError(
            "discount_parser: no holdings parsed — check file format or header detection"
        )

    invested_ils = sum(h.market_value_ils for h in holdings)
    cash_ils = total_value_ils - invested_ils

    return PortfolioSnapshot(
        snapshot_time=datetime.now(),
        total_value_ils=total_value_ils,
        cash_ils=max(cash_ils, Decimal("0")),  # guard against rounding below zero
        invested_ils=invested_ils,
        day_pnl_ils=day_pnl_ils,
        day_pnl_pct=day_pnl_pct,
        holdings=holdings,
    )
